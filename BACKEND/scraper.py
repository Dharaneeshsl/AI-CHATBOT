import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import os
from urllib.parse import urljoin, urlparse
import PyPDF2
import docx
from openpyxl import load_workbook
from collections import deque

class MOSDACScraper:
    def __init__(self, base_url, output_dir="extracted_content", max_depth=3):
        self.base_url = base_url
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.visited_urls = set()
        self.to_visit_queue = deque()
        self.allowed_domains = [urlparse(base_url).netloc]
        self.max_depth = max_depth
        self.file_extensions = [".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".zip", ".rar", ".tar", ".jar"]

    def _is_valid_url(self, url, current_depth):
        parsed_url = urlparse(url)
        if parsed_url.netloc not in self.allowed_domains:
            return False
        if any(path_segment in parsed_url.path for path_segment in ["login", "signup", "auth", "reset-credentials", "image", "img", "css", "js"]):
            return False
        if url.split("#")[0] in self.visited_urls:
            return False
        if current_depth > self.max_depth:
            return False
        return True

    async def _download_file(self, page, url, save_path):
        try:
            response = await page.request.get(url)
            if response.ok:
                with open(save_path, 'wb') as f:
                    f.write(await response.body())
                print(f"Downloaded: {url} to {save_path}")
                return True
            else:
                print(f"Failed to download {url}: Status {response.status}")
                return False
        except Exception as e:
            print(f"Error downloading {url}: {e}")
            return False

    def _extract_text_from_pdf(self, pdf_path):
        text = ""
        try:
            with open(pdf_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page_num in range(len(reader.pages)):
                    text += reader.pages[page_num].extract_text()
            return text
        except Exception as e:
            print(f"Error extracting text from PDF {pdf_path}: {e}")
            return None

    def _extract_text_from_docx(self, docx_path):
        text = []
        try:
            document = docx.Document(docx_path)
            for paragraph in document.paragraphs:
                text.append(paragraph.text)
            return '\n'.join(text)
        except Exception as e:
            print(f"Error extracting text from DOCX {docx_path}: {e}")
            return None

    def _extract_text_from_xlsx(self, xlsx_path):
        text = []
        try:
            workbook = load_workbook(xlsx_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                    text.append('\t'.join(row_data))
            return '\n'.join(text)
        except Exception as e:
            print(f"Error extracting text from XLSX {xlsx_path}: {e}")
            return None

    async def _process_page(self, page, url, current_depth):
        await page.wait_for_load_state("networkidle")
        soup = BeautifulSoup(await page.content(), 'html.parser')

        page_text = ' '.join(p.get_text() for p in soup.find_all('p'))
        title = await page.title()
        
        filename_base = urlparse(url).path.replace('/', '_').strip('_')
        if not filename_base:
            filename_base = 'index'
        filename_base = filename_base.split('?')[0].split('#')[0]

        html_filename = os.path.join(self.output_dir, f"{filename_base}.html")
        with open(html_filename, 'w', encoding='utf-8') as f:
            f.write(soup.prettify())
        print(f"Saved HTML: {url} to {html_filename}")

        text_filename = os.path.join(self.output_dir, f"{filename_base}.txt")
        with open(text_filename, 'w', encoding='utf-8') as f:
            f.write(f"Title: {title}\n\n")
            f.write(page_text)
        print(f"Extracted text: {url} to {text_filename}")

        tables = soup.find_all('table')
        if tables:
            table_filename = os.path.join(self.output_dir, f"{filename_base}_tables.txt")
            with open(table_filename, 'w', encoding='utf-8') as f:
                for i, table in enumerate(tables):
                    f.write(f"\n--- Table {i+1} ---\n")
                    for row in table.find_all('tr'):
                        row_data = [cell.get_text(strip=True) for cell in row.find_all(['td', 'th'])]
                        f.write('\t'.join(row_data) + '\n')
            print(f"Extracted tables from {url} to {table_filename}")

        for link in soup.find_all('a', href=True):
            href = link['href']
            full_url = urljoin(url, href)
            normalized_url = full_url.split('?')[0].split('#')[0]

            if self._is_valid_url(normalized_url, current_depth + 1):
                self.to_visit_queue.append((normalized_url, current_depth + 1))

    async def scrape(self, start_url):
        self.to_visit_queue.append((start_url, 0))

        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()

            while self.to_visit_queue:
                url, current_depth = self.to_visit_queue.popleft()
                
                if url.split("#")[0] in self.visited_urls:
                    continue
                self.visited_urls.add(url.split("#")[0])

                print(f"Scraping: {url} (Depth: {current_depth})")

                try:
                    # Check if it's a file to download directly
                    is_file = False
                    for ext in self.file_extensions:
                        if url.endswith(ext):
                            file_name = os.path.basename(urlparse(url).path)
                            file_path = os.path.join(self.output_dir, file_name)
                            if await self._download_file(page, url, file_path):
                                if ext == ".pdf":
                                    extracted_text = self._extract_text_from_pdf(file_path)
                                elif ext == ".docx":
                                    extracted_text = self._extract_text_from_docx(file_path)
                                elif ext == ".xlsx":
                                    extracted_text = self._extract_text_from_xlsx(file_path)
                                else:
                                    extracted_text = None

                                if extracted_text:
                                    text_output_path = os.path.join(self.output_dir, f"{file_name}.txt")
                                    with open(text_output_path, 'w', encoding='utf-8') as f:
                                        f.write(extracted_text)
                                    print(f"Extracted text from {file_name} to {text_output_path}")
                            is_file = True
                            break
                    
                    if not is_file:
                        await page.goto(url, wait_until='networkidle')
                        await self._process_page(page, url, current_depth)

                except Exception as e:
                    print(f"Error processing {url}: {e}")
            
            await browser.close()

async def main():
    scraper = MOSDACScraper("https://www.mosdac.gov.in", max_depth=2) # Limiting depth for initial run
    await scraper.scrape("https://www.mosdac.gov.in")

if __name__ == "__main__":
    asyncio.run(main())

